"""
Exporta informe de usuarios y estado de cambio de contraseña a Excel.
Uso: python exportar_usuarios_contrasena.py
Requiere: psycopg2-binary, openpyxl
"""
import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Conexión ──────────────────────────────────────────────────────────────────
DB = {
    "host":     "database-1.chgqoo4oaal4.us-east-2.rds.amazonaws.com",
    "port":     5432,
    "dbname":   "contabilidadcq",
    "user":     "postgres",
    "password": "Samuel22.",
    "sslmode":  "require",
}

QUERY = """
SELECT
    u.nombre                                        AS "Nombre",
    u.email                                         AS "Email",
    r.nombre                                        AS "Rol",
    COALESCE(a.nombre, '—')                         AS "Área",
    CASE WHEN u.is_active THEN 'Activo' ELSE 'Inactivo' END AS "Estado cuenta",
    CASE
        WHEN u.must_change_password THEN 'PENDIENTE'
        ELSE 'Cambiada'
    END                                             AS "Contraseña"
FROM users u
LEFT JOIN roles r ON u.role_id = r.id
LEFT JOIN areas a ON u.area_id  = a.id
ORDER BY u.must_change_password DESC, u.nombre;
"""

# ── Estilos ───────────────────────────────────────────────────────────────────
COLOR_HEADER    = "1F4E79"   # azul oscuro
COLOR_PENDIENTE = "FFD7D7"   # rojo claro
COLOR_CAMBIADA  = "D7F3D7"   # verde claro
COLOR_FILA_PAR  = "EBF3FB"   # azul muy claro

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def aplicar_header(ws, columnas):
    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor=COLOR_HEADER)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, titulo in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col_idx, value=titulo)
        celda.font      = header_font
        celda.fill      = header_fill
        celda.alignment = header_align
        celda.border    = thin_border()
    ws.row_dimensions[1].height = 28

def aplicar_filas(ws, filas, col_contrasena_idx):
    fill_pendiente = PatternFill("solid", fgColor=COLOR_PENDIENTE)
    fill_cambiada  = PatternFill("solid", fgColor=COLOR_CAMBIADA)
    fill_par       = PatternFill("solid", fgColor=COLOR_FILA_PAR)
    font_normal    = Font(name="Calibri", size=10)
    font_pendiente = Font(name="Calibri", size=10, bold=True, color="C00000")
    font_cambiada  = Font(name="Calibri", size=10, color="375623")

    for row_idx, fila in enumerate(filas, start=2):
        es_pendiente = str(fila[col_contrasena_idx - 1]) == "PENDIENTE"
        row_fill = fill_pendiente if es_pendiente else (fill_par if row_idx % 2 == 0 else None)

        for col_idx, valor in enumerate(fila, start=1):
            celda = ws.cell(row=row_idx, column=col_idx, value=valor)
            celda.border    = thin_border()
            celda.alignment = Alignment(vertical="center")

            if col_idx == col_contrasena_idx:
                celda.fill = fill_pendiente if es_pendiente else fill_cambiada
                celda.font = font_pendiente if es_pendiente else font_cambiada
                celda.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if row_fill:
                    celda.fill = row_fill
                celda.font = font_normal

        ws.row_dimensions[row_idx].height = 18

def ajustar_columnas(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

def agregar_resumen(wb, total, pendientes, cambiadas):
    ws = wb.create_sheet("Resumen")
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)

    titulo = ws.cell(row=1, column=1, value="Resumen — Estado de contraseñas")
    titulo.font = Font(name="Calibri", bold=True, size=14, color=COLOR_HEADER)
    ws.merge_cells("A1:C1")
    titulo.alignment = Alignment(horizontal="center")

    ws.cell(row=3, column=1, value="Estado").font  = header_font
    ws.cell(row=3, column=1).fill = header_fill
    ws.cell(row=3, column=2, value="Cantidad").font = header_font
    ws.cell(row=3, column=2).fill = header_fill
    ws.cell(row=3, column=3, value="% del total").font = header_font
    ws.cell(row=3, column=3).fill = header_fill

    datos = [
        ("PENDIENTE (no han cambiado)", pendientes, pendientes / total if total else 0),
        ("Contraseña cambiada",          cambiadas,  cambiadas  / total if total else 0),
        ("Total usuarios",               total,      1.0),
    ]
    fills = [
        PatternFill("solid", fgColor=COLOR_PENDIENTE),
        PatternFill("solid", fgColor=COLOR_CAMBIADA),
        PatternFill("solid", fgColor="E2EFDA"),
    ]
    for i, (estado, cant, pct) in enumerate(datos, start=4):
        ws.cell(row=i, column=1, value=estado).font  = Font(name="Calibri", size=11)
        ws.cell(row=i, column=2, value=cant).font    = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=i, column=3, value=pct).number_format = "0.0%"
        ws.cell(row=i, column=3).font = Font(name="Calibri", size=11)
        for c in range(1, 4):
            ws.cell(row=i, column=c).fill   = fills[i - 4]
            ws.cell(row=i, column=c).border = thin_border()
            ws.cell(row=i, column=c).alignment = Alignment(vertical="center")
        ws.row_dimensions[i].height = 20

    for col, ancho in [("A", 35), ("B", 15), ("C", 15)]:
        ws.column_dimensions[col].width = ancho

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("Conectando a la base de datos...")
    conn = psycopg2.connect(**DB)
    cur  = conn.cursor()
    cur.execute(QUERY)
    filas    = cur.fetchall()
    columnas = [desc[0] for desc in cur.description]
    cur.close()
    conn.close()
    print(f"  {len(filas)} usuarios encontrados.")

    col_contrasena_idx = columnas.index("Contraseña") + 1
    pendientes = sum(1 for f in filas if f[col_contrasena_idx - 1] == "PENDIENTE")
    cambiadas  = len(filas) - pendientes

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    ws.sheet_view.showGridLines = False

    aplicar_header(ws, columnas)
    aplicar_filas(ws, filas, col_contrasena_idx)
    ajustar_columnas(ws)
    ws.freeze_panes = "A2"

    agregar_resumen(wb, len(filas), pendientes, cambiadas)

    fecha    = datetime.now().strftime("%Y%m%d_%H%M")
    archivo  = f"informe_contrasenas_{fecha}.xlsx"
    wb.save(archivo)
    print(f"\nArchivo generado: {archivo}")
    print(f"  Pendientes : {pendientes}")
    print(f"  Cambiadas  : {cambiadas}")
    print(f"  Total      : {len(filas)}")

if __name__ == "__main__":
    main()
