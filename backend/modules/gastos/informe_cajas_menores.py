# -*- coding: utf-8 -*-
"""Informe Docuflow de legalizaciones de cajas menores.

Espejo de informe_zonas.py pero agrupado por caja menor. Cubre los flujos de
reembolso ('mantenimiento' y 'general'): los técnicos de mantenimiento llevan
su caja en el nombre de usuario ("Nombre - TESORERIA N") y de ahí se extrae;
si el nombre no la trae, se usa el área del paquete. El corte incluye los
paquetes cuya semana INICIA en o antes de fecha_hasta, con la semana completa;
fecha_desde (opcional) acota por el inicio de la semana.
"""
import io
import re
from datetime import date
from typing import Optional
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from db.models import (
    PaqueteGasto, GastoLegalizacion, User, Area,
    CuentaAuxiliar, CentroCosto, CentroOperacion,
)

BOGOTA = ZoneInfo("America/Bogota")
EMAIL_PRUEBA = "tecnico@cafequindio.com"

# Caja menor embebida en el nombre del usuario: "Nombre - TESORERIA N"
# (a veces sin espacios: "Nombre -TESORERIA 12"). Si no aparece, se usa el
# área del paquete como caja.
_RE_CAJA = re.compile(r"TESORER[ÍI]A\s*(\d+)", re.IGNORECASE)


def _caja_de(nombre: Optional[str], area: Optional[str]) -> str:
    m = _RE_CAJA.search(nombre or "")
    if m:
        return f"TESORERIA {m.group(1)}"
    return area or "Sin área"


ESTADOS = {
    "borrador": "Borrador", "en_validacion": "En validación",
    "en_revision": "En revisión", "devuelto": "Devuelto",
    "aprobado": "Aprobado", "en_tesoreria": "En Tesorería",
    "pagado": "Pagado", "cruzado": "Cruzado",
}
MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
         "agosto", "septiembre", "octubre", "noviembre", "diciembre"]


def _fmt_ts(ts) -> Optional[str]:
    if ts is None:
        return None
    return ts.astimezone(BOGOTA).strftime("%Y-%m-%d %H:%M")


def _paquete_filters(q, fecha_desde: Optional[date], fecha_hasta: date):
    q = (q.where(PaqueteGasto.tipo_flujo.in_(("mantenimiento", "general")))
          .where(User.email != EMAIL_PRUEBA)
          .where(PaqueteGasto.fecha_inicio <= fecha_hasta))
    if fecha_desde is not None:
        q = q.where(PaqueteGasto.fecha_inicio >= fecha_desde)
    return q


async def consultar_datos(
    db: AsyncSession, fecha_desde: Optional[date], fecha_hasta: date
) -> tuple[list[dict], list[dict]]:
    """Trae paquetes y detalle como listas de dicts (columnas planas, sin ORM cascada)."""
    q_paq = _paquete_filters(
        select(
            User.email, User.nombre.label("usuario"),
            Area.nombre.label("area"), PaqueteGasto.folio, PaqueteGasto.tipo_flujo,
            PaqueteGasto.semana, PaqueteGasto.fecha_inicio, PaqueteGasto.fecha_fin,
            PaqueteGasto.estado, PaqueteGasto.total_documentos,
            PaqueteGasto.monto_total, PaqueteGasto.monto_a_pagar,
            PaqueteGasto.fecha_envio, PaqueteGasto.fecha_aprobacion,
            PaqueteGasto.fecha_envio_tesoreria, PaqueteGasto.fecha_pago,
            PaqueteGasto.fecha_cruce,
        )
        .join(User, User.id == PaqueteGasto.user_id)
        .outerjoin(Area, Area.id == PaqueteGasto.area_id),
        fecha_desde, fecha_hasta,
    ).order_by(Area.nombre, User.email, PaqueteGasto.semana, PaqueteGasto.folio)

    q_det = _paquete_filters(
        select(
            User.email, User.nombre.label("usuario"),
            Area.nombre.label("area"), PaqueteGasto.folio, PaqueteGasto.tipo_flujo,
            PaqueteGasto.semana, PaqueteGasto.estado.label("estado_paquete"),
            GastoLegalizacion.fecha, GastoLegalizacion.no_identificacion,
            GastoLegalizacion.pagado_a, GastoLegalizacion.concepto,
            GastoLegalizacion.no_recibo, GastoLegalizacion.valor_pagado,
            GastoLegalizacion.estado_gasto,
            CuentaAuxiliar.codigo.label("cod_cuenta_auxiliar"),
            CuentaAuxiliar.descripcion.label("cuenta_auxiliar"),
            CentroCosto.nombre.label("centro_costo"),
            CentroOperacion.nombre.label("centro_operacion"),
        )
        .select_from(GastoLegalizacion)
        .join(PaqueteGasto, PaqueteGasto.id == GastoLegalizacion.paquete_id)
        .join(User, User.id == PaqueteGasto.user_id)
        .outerjoin(Area, Area.id == PaqueteGasto.area_id)
        .outerjoin(CuentaAuxiliar, CuentaAuxiliar.id == GastoLegalizacion.cuenta_auxiliar_id)
        .outerjoin(CentroCosto, CentroCosto.id == GastoLegalizacion.centro_costo_id)
        .outerjoin(CentroOperacion, CentroOperacion.id == GastoLegalizacion.centro_operacion_id),
        fecha_desde, fecha_hasta,
    ).order_by(Area.nombre, User.email, PaqueteGasto.semana, PaqueteGasto.folio, GastoLegalizacion.orden)

    paq = [dict(r._mapping) for r in (await db.execute(q_paq)).all()]
    det = [dict(r._mapping) for r in (await db.execute(q_det)).all()]
    return paq, det


def construir_excel(
    paq: list[dict], det: list[dict],
    fecha_desde: Optional[date], fecha_hasta: date,
) -> bytes:
    """Arma el workbook de 4 hojas. Síncrono: llamar con asyncio.to_thread."""
    FLUJOS = {"mantenimiento": "Mantenimiento", "general": "General"}
    for p in paq:
        p["area"] = _caja_de(p["usuario"], p["area"])
        p["tipo_flujo"] = FLUJOS.get(p["tipo_flujo"], p["tipo_flujo"])
        p["monto_total"] = float(p["monto_total"] or 0)
        p["monto_a_pagar"] = float(p["monto_a_pagar"]) if p["monto_a_pagar"] is not None else None
    for g in det:
        g["area"] = _caja_de(g["usuario"], g["area"])
        g["tipo_flujo"] = FLUJOS.get(g["tipo_flujo"], g["tipo_flujo"])
        g["valor_pagado"] = float(g["valor_pagado"] or 0)

    rep = [p for p in paq if p["estado"] != "borrador"]

    montos_area: dict[str, float] = {}
    for p in rep:
        montos_area[p["area"]] = montos_area.get(p["area"], 0) + p["monto_total"]
    orden_areas = sorted(montos_area, key=lambda a: -montos_area[a])
    idx_area = {a: i for i, a in enumerate(orden_areas)}

    desde_txt = (f"del {fecha_desde.day} de {MESES[fecha_desde.month - 1]} de {fecha_desde.year} "
                 if fecha_desde else "desde el inicio ")
    corte_txt = (f"Corte: {desde_txt}al {fecha_hasta.day} de "
                 f"{MESES[fecha_hasta.month - 1]} de {fecha_hasta.year}"
                 " — Fuente: Docuflow (producción)")

    # ---- Resumen por Área (caja menor) ----
    def _resumen_area(area: Optional[str]):
        r = rep if area is None else [p for p in rep if p["area"] == area]
        pagado = sum(p["monto_total"] for p in r if p["estado"] in ("pagado", "cruzado"))
        total = sum(p["monto_total"] for p in r)
        return [area or "TOTAL GENERAL", len({p["email"] for p in r}), len(r),
                sum(p["total_documentos"] for p in r), total, pagado, total - pagado]

    area_rows = [_resumen_area(a) for a in orden_areas] + [_resumen_area(None)]
    area_hdr = ["Caja menor", "Usuarios", "Paquetes reportados", "Documentos",
                "Monto reportado", "Monto pagado", "Monto en trámite"]

    # ---- Resumen por Usuario ----
    por_usuario: dict[tuple, list[dict]] = {}
    for p in rep:
        por_usuario.setdefault((p["area"], p["email"]), []).append(p)
    usr_rows = []
    for (area, email), r in por_usuario.items():
        cnt = lambda e: sum(1 for p in r if p["estado"] == e)  # noqa: E731
        usr_rows.append([
            area, r[0]["usuario"], email, len(r),
            sum(p["total_documentos"] for p in r),
            sum(p["monto_total"] for p in r),
            min(p["semana"] for p in r), max(p["semana"] for p in r),
            cnt("en_revision"), cnt("devuelto"), cnt("aprobado"),
            cnt("en_tesoreria"), cnt("pagado"), cnt("cruzado"),
        ])
    usr_rows.sort(key=lambda x: (idx_area.get(x[0], 99), -x[5]))
    usr_hdr = ["Caja menor", "Usuario", "Email", "Paquetes", "Documentos", "Monto reportado",
               "Primera semana", "Última semana", "En revisión", "Devuelto", "Aprobado",
               "En Tesorería", "Pagado", "Cruzado"]

    # ---- Paquetes ----
    paq_s = sorted(paq, key=lambda p: (idx_area.get(p["area"], 99), p["usuario"], p["semana"]))
    paq_hdr = ["Caja menor", "Usuario", "Folio", "Flujo", "Semana", "Inicio", "Fin", "Estado", "Docs",
               "Monto total", "Monto a pagar", "Fecha envío", "Fecha aprobación",
               "Envío tesorería", "Fecha pago"]
    paq_rows = [[
        p["area"], p["usuario"], p["folio"], p["tipo_flujo"], p["semana"],
        str(p["fecha_inicio"]), str(p["fecha_fin"]),
        ESTADOS.get(p["estado"], p["estado"]), p["total_documentos"],
        p["monto_total"], p["monto_a_pagar"],
        _fmt_ts(p["fecha_envio"]), _fmt_ts(p["fecha_aprobacion"]),
        _fmt_ts(p["fecha_envio_tesoreria"]), _fmt_ts(p["fecha_pago"]),
    ] for p in paq_s]

    # ---- Detalle Gastos ----
    det_s = sorted(det, key=lambda g: (idx_area.get(g["area"], 99), g["usuario"],
                                       g["semana"], g["folio"] or ""))
    det_hdr = ["Caja menor", "Usuario", "Folio paquete", "Semana", "Estado paquete", "Fecha gasto",
               "NIT/CC", "Pagado a", "Concepto", "No. recibo", "Valor pagado", "Estado gasto",
               "Cód. cuenta auxiliar", "Cuenta auxiliar", "Centro de costo", "Centro de operación"]
    det_rows = [[
        g["area"], g["usuario"], g["folio"], g["semana"],
        ESTADOS.get(g["estado_paquete"], g["estado_paquete"]), str(g["fecha"]),
        g["no_identificacion"], g["pagado_a"], g["concepto"], g["no_recibo"],
        g["valor_pagado"], g["estado_gasto"], g["cod_cuenta_auxiliar"],
        g["cuenta_auxiliar"], g["centro_costo"], g["centro_operacion"],
    ] for g in det_s]

    # ---- Escritura con formato (mismo estilo del informe de zonas) ----
    wb = Workbook()
    AZUL = "1F4E5F"
    GRIS = "F2F2F2"
    thin = Border(bottom=Side(style="thin", color="D9D9D9"))

    def hoja(ws, titulo, hdr, rows, money_cols, total_row=False):
        ws.append([titulo]); ws.append([corte_txt]); ws.append([])
        ws["A1"].font = Font(bold=True, size=13, color=AZUL)
        ws["A2"].font = Font(italic=True, size=9, color="666666")
        ws.append(hdr)
        hr = ws.max_row
        for c in range(1, len(hdr) + 1):
            cell = ws.cell(row=hr, column=c)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor=AZUL)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for i, row in enumerate(rows):
            ws.append(row)
            r = ws.max_row
            for c in range(1, len(hdr) + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin
                cell.font = Font(size=10)
                if i % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor=GRIS)
                if c in money_cols:
                    cell.number_format = "#,##0"
            if total_row and i == len(rows) - 1:
                for c in range(1, len(hdr) + 1):
                    ws.cell(row=r, column=c).font = Font(bold=True, size=10)
        for c in range(1, len(hdr) + 1):
            w = max([len(str(hdr[c - 1]))] +
                    [len(str(row[c - 1])) for row in rows if row[c - 1] is not None][:200])
            ws.column_dimensions[get_column_letter(c)].width = min(max(w + 2, 10), 42)
        ws.freeze_panes = ws.cell(row=hr + 1, column=1)

    ws = wb.active; ws.title = "Resumen por Caja Menor"
    hoja(ws, "INFORME DOCUFLOW — LEGALIZACIONES DE CAJAS MENORES",
         area_hdr, area_rows, {5, 6, 7}, total_row=True)
    hoja(wb.create_sheet("Resumen por Usuario"),
         "Resumen por usuario (solo paquetes reportados, excluye borradores)",
         usr_hdr, usr_rows, {6})
    hoja(wb.create_sheet("Paquetes"),
         "Listado completo de paquetes (incluye borradores no enviados)",
         paq_hdr, paq_rows, {10, 11})
    hoja(wb.create_sheet("Detalle Gastos"),
         "Detalle de todos los gastos legalizados línea a línea",
         det_hdr, det_rows, {11})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
